#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


MASTER_OPERATIONAL_COLUMNS = [
    "最近分配日期",
    "负责人名",
    "是否回复",
    "实际回复类型",
    "回复内容",
    "回复日期",
    "跟进状态",
    "人工备注",
    "来源回复分配表",
]


def normalize_url(url):
    if not url:
        return ""
    url = str(url).strip()
    if url.startswith("/"):
        url = "https://www.reddit.com" + url
    return url.rstrip("/")


def parse_date_from_name(path):
    match = re.search(r"(20\d{2}-\d{2}-\d{2})", Path(path).name)
    if match:
        return match.group(1)
    return datetime.now().strftime("%Y-%m-%d")


def header_map(ws):
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def ensure_column(ws, headers, name):
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=name)
    headers[name] = col
    return col


def first_existing(headers, names):
    for name in names:
        if name in headers:
            return headers[name]
    return None


def get_value(ws, row, col):
    if not col:
        return ""
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


def set_if_meaningful(ws, row, col, value, preserve_done=False):
    value = "" if value is None else str(value).strip()
    if not value:
        return False
    old = get_value(ws, row, col)
    if preserve_done and old in {"是", "不适合回复", "帖子回复", "评论回复", "both", "不回复", "已回复", "不适合回复"}:
        if value in {"否", "未回复", "待回复"}:
            return False
    if old == value:
        return False
    ws.cell(row=row, column=col, value=value)
    return True


def choose_master_sheet(wb):
    for name in ["Master Posts", "全量数据库", "Posts"]:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def collect_assignment_rows(wb):
    rows = []
    for sheet_name in ["待回复", "已处理"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = header_map(ws)
        url_col = first_existing(headers, ["原帖链接", "帖子链接", "Post URL"])
        title_col = first_existing(headers, ["帖子标题", "Post Title"])
        if not url_col:
            continue
        for row_idx in range(2, ws.max_row + 1):
            url = normalize_url(get_value(ws, row_idx, url_col))
            if not url:
                continue
            item = {
                "source_sheet": sheet_name,
                "原帖链接": url,
                "帖子标题": get_value(ws, row_idx, title_col),
            }
            for field in ["负责人名", "是否回复", "实际回复类型", "回复内容", "备注", "回复日期", "跟进状态"]:
                col = headers.get(field)
                item[field] = get_value(ws, row_idx, col)
            rows.append(item)
    return rows


def derive_status(item, old_status):
    status = item.get("跟进状态") or ""
    if status:
        return status
    replied = item.get("是否回复") or ""
    reply_type = item.get("实际回复类型") or ""
    if replied == "是" or reply_type in {"帖子回复", "评论回复", "both"}:
        return "已回复"
    if replied == "不适合回复" or reply_type == "不回复":
        return "不适合回复"
    if replied == "待复核":
        return "待人工复核"
    if old_status in {"已回复", "不适合回复", "已归档"}:
        return old_status
    return "待回复"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--assignment", required=True, help="Reply assignment workbook with 待回复 sheet.")
    parser.add_argument("--master", required=True, help="Existing master workbook to update.")
    parser.add_argument("--output", required=True, help="Updated master workbook output path.")
    parser.add_argument("--assignment-date", help="Date to write into 最近分配日期. Defaults to date in assignment filename or today.")
    args = parser.parse_args()

    assignment_date = args.assignment_date or parse_date_from_name(args.assignment)
    source_name = Path(args.assignment).name

    assignment_wb = load_workbook(args.assignment, data_only=True)
    assignment_rows = collect_assignment_rows(assignment_wb)

    master_wb = load_workbook(args.master)
    master_ws = choose_master_sheet(master_wb)
    headers = header_map(master_ws)
    url_col = first_existing(headers, ["原帖链接", "帖子链接", "Post URL"])
    title_col = first_existing(headers, ["帖子标题", "Post Title"])
    if not url_col:
        raise SystemExit("Master workbook is missing a URL column: expected 原帖链接 / 帖子链接 / Post URL")

    for column in MASTER_OPERATIONAL_COLUMNS:
        ensure_column(master_ws, headers, column)

    url_to_row = {}
    for row_idx in range(2, master_ws.max_row + 1):
        url = normalize_url(get_value(master_ws, row_idx, url_col))
        if url:
            url_to_row[url] = row_idx

    updated = 0
    appended = 0
    for item in assignment_rows:
        url = item["原帖链接"]
        row_idx = url_to_row.get(url)
        if not row_idx:
            row_idx = master_ws.max_row + 1
            master_ws.cell(row=row_idx, column=url_col, value=url)
            if title_col:
                master_ws.cell(row=row_idx, column=title_col, value=item.get("帖子标题", ""))
            url_to_row[url] = row_idx
            appended += 1

        changed = False
        changed |= set_if_meaningful(master_ws, row_idx, headers["最近分配日期"], assignment_date)
        changed |= set_if_meaningful(master_ws, row_idx, headers["来源回复分配表"], source_name)
        changed |= set_if_meaningful(master_ws, row_idx, headers["负责人名"], item.get("负责人名"))
        changed |= set_if_meaningful(master_ws, row_idx, headers["是否回复"], item.get("是否回复"), preserve_done=True)
        changed |= set_if_meaningful(master_ws, row_idx, headers["实际回复类型"], item.get("实际回复类型"), preserve_done=True)
        changed |= set_if_meaningful(master_ws, row_idx, headers["回复内容"], item.get("回复内容"))
        changed |= set_if_meaningful(master_ws, row_idx, headers["人工备注"], item.get("备注"))

        old_status = get_value(master_ws, row_idx, headers["跟进状态"])
        status = derive_status(item, old_status)
        changed |= set_if_meaningful(master_ws, row_idx, headers["跟进状态"], status, preserve_done=True)

        if status == "已回复":
            reply_date = item.get("回复日期") or assignment_date
            changed |= set_if_meaningful(master_ws, row_idx, headers["回复日期"], reply_date)

        if changed:
            updated += 1

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(out)
    print(json.dumps({
        "output": str(out),
        "assignment_rows": len(assignment_rows),
        "updated_rows": updated,
        "appended_rows": appended,
        "assignment_date": assignment_date,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
