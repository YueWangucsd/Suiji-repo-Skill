#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUBREDDIT_GROUPS = {
    "jewelry": "首饰主社区",
    "Earrings": "耳饰社区",
    "jewelrymaking": "首饰制作/工艺",
    "jewelers": "珠宝专业人士",
    "JewelryDesign": "首饰设计验证",
    "jewellerymaking": "首饰制作/工艺",
    "femalefashionadvice": "穿搭/职场",
    "BusinessFashion": "穿搭/职场",
    "fashionwomens35": "成熟穿搭/品质",
    "OUTFITS": "穿搭反馈",
    "PetiteFashionAdvice": "小个子穿搭",
    "Weddingattireapproval": "婚礼/场景穿搭",
    "piercing": "穿孔/敏感耳痛点",
    "PiercingAdvice": "穿孔建议/痛点",
    "SkincareAddiction": "皮肤敏感相关",
    "BuyItForLife": "耐用品/质量信任",
    "capsulewardrobe": "胶囊衣橱/少而精",
    "minimalism": "极简/少而精",
    "HerOneBag": "女性旅行轻量化",
    "onebag": "旅行轻量化",
}


TAG_PATTERNS = {
    "痛点：掉色/氧化/镀层磨损": r"\btarnish|fading|plating|plated|oxidation|discolor|green skin|turn(?:ing)? my skin green|rust|vermeil|change colors?",
    "痛点：敏感/刺激/发炎": r"sensitive ears?|nickel|hypoallergenic|irritation|allergic|contact dermatitis|infect|swollen|itch|bioflex|keloid|bump",
    "痛点：舒适度/重量": r"heavy earrings?|too heavy|hurt|uncomfortable|weight|flat back|threadless|sleep|ripping|sagging",
    "痛点：损坏/维修/拆装困难": r"broken|broke|clasp|loose|stone fell|repair|fix|snapped|fell out|stuck|can't pull|cannot pull",
    "痛点：造型/职场场景": r"office|work appropriate|business casual|workwear|professional|too much|date night|formal|wedding|what jewelry|which jewelry|choose between",
    "痛点：旅行/收纳/缠绕": r"travel|packing|storage|tangled|organizer|lost earring|lost jewelry|trip|onebag|capsule",
    "痛点：价格/质量/信任": r"worth it|overpriced|avoid|dupe|recommendation|buyer beware|cheap|bad quality|poor quality|review|scam|price for|good price|retains value|insure|insurance",
    "机会：模块化/可转换": r"modular|convertible|detachable|interchangeable|versatile|multi-use|day to night|threadless|capsule",
}

INTENT_PATTERNS = {
    "意图：抱怨/负面体验": r"regret|warning|beware|problem|issue|broke|broken|fell out|can't|cannot|help|struggling|scam|disappointed|complaint",
    "意图：求建议/求解决方案": r"\?|advice|recommend|suggest|what should|how do|how to|looking for|help me|thoughts",
    "意图：购买研究/品牌比较": r"buy|purchase|price|worth it|recommendation|brand|where are we buying|find|dupe",
    "意图：保养/维修/护理": r"care|clean|polish|repair|fix|sealant|tarnish|untangle",
    "意图：展示/晒图": r"show|collection|fit|outfit|made|work in progress|proud",
}

PRODUCT_PATTERNS = {
    "产品相关：钛合金/低敏材质": r"titanium|sensitive|nickel|hypoallergenic|irritation|bioflex|skin green|green skin",
    "产品相关：Core-Lock/快拆结构": r"detachable|threadless|clicker|clasp|stuck|fell out|loose|can't pull|interchangeable|change",
    "产品相关：职场到社交切换": r"office|work|professional|business casual|evening|date|formal|wedding",
    "产品相关：旅行/胶囊衣橱": r"travel|packing|capsule|onebag|versatile|multi-use|storage|tangled",
    "产品相关：质量/价值证明": r"quality|worth it|price|overpriced|durable|long lasting|bifl|review|brand",
}

PRODUCT_DESIGN_PATTERNS = {
    "重量": r"heavy earrings?|too heavy|weight|weighs?|weighted|sagging|ripping|pulling|earlobe|lobe stretching",
    "过敏/敏感": r"sensitive ears?|nickel|hypoallergenic|allerg(?:y|ic)|irritation|contact dermatitis|itch|swollen|infect|bump|keloid|bioflex|skin green|green skin",
    "佩戴舒适度": r"uncomfortable|comfort|hurt|hurts|pain|painful|sleep|flat back|threadless|backs?|post length|too tight|too loose|rubbing|pinching",
}

JEWELRY_TERMS = r"jewelry|jewellery|earring|earrings|ring|rings|necklace|bracelet|pendant|chain|pearl|silver|gold|plated|tarnish|vermeil|sterling|gem|stone|ruby|diamond|piercing|piercings|lobe|lobes|helix|conch|rook|daith|tragus|barbell|labret|stud|hoop|clasp|bead|jump ring|wire|titanium"

JEWELRY_TYPE_PATTERNS = {
    "耳饰": r"earring|earrings|ear cuff|stud|hoop|huggie|drop earring|dangle|ear jacket",
    "戒指": r"\bring\b|rings|stacking ring|wedding band|engagement ring",
    "项链/吊坠": r"necklace|pendant|chain|choker",
    "手链/手镯": r"bracelet|bangle|cuff",
    "穿孔饰品": r"piercing|piercings|lobe|lobes|helix|conch|rook|daith|tragus|barbell|labret|flat back|threadless|post",
    "扣件/结构": r"clasp|lock|clicker|threadless|detachable|interchangeable|modular|convertible|stuck|loose|fell out|lost backing|backing",
    "材质/过敏": r"titanium|nickel|hypoallergenic|allerg(?:y|ic)|sensitive|irritation|gold|silver|plated|vermeil|stainless steel|green skin|tarnish",
    "佩戴舒适度": r"comfort|uncomfortable|hurt|hurts|pain|painful|heavy|weight|sleep|too tight|too loose|pinching|rubbing",
    "收纳/旅行": r"travel|packing|storage|tangled|organizer|capsule|onebag",
    "造型/职场": r"office|work|professional|business casual|formal|wedding|outfit|style|styling",
}

CORE_JEWELRY_SUBS = {
    "jewelry",
    "Earrings",
    "jewelrymaking",
    "jewelers",
    "JewelryDesign",
    "jewellerymaking",
    "piercing",
    "PiercingAdvice",
}


def match_labels(text, patterns):
    return [label for label, pattern in patterns.items() if re.search(pattern, text, re.I)]


def joined(labels):
    return ", ".join(labels)


def compact_count(value):
    if value in ("", None):
        return ""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return ""
    if value >= 1_000_000:
        return f"{value / 1_000_000:.1f}M"
    if value >= 1_000:
        return f"{value / 1_000:.0f}K"
    return str(int(value))


def size_band(subscribers):
    if subscribers in ("", None):
        return "未知"
    try:
        subscribers = int(subscribers)
    except (TypeError, ValueError):
        return "未知"
    if subscribers >= 1_000_000:
        return "100万+"
    if subscribers >= 500_000:
        return "50万-100万"
    if subscribers >= 100_000:
        return "10万-50万"
    if subscribers >= 50_000:
        return "5万-10万"
    if subscribers >= 10_000:
        return "1万-5万"
    return "1万以下"


def active_ratio(active, subscribers):
    try:
        active = int(active)
        subscribers = int(subscribers or 0)
    except (TypeError, ValueError):
        return ""
    if not subscribers:
        return ""
    return round(active / subscribers, 4)


def activity_band(active, subscribers):
    if active in ("", None):
        return "未知"
    try:
        active = int(active)
        subscribers = int(subscribers or 0)
    except (TypeError, ValueError):
        return "未知"
    ratio = active / subscribers if subscribers else 0
    if active >= 10_000 or ratio >= 0.01:
        return "高活跃"
    if active >= 1_000 or ratio >= 0.003:
        return "中活跃"
    if active > 0:
        return "低活跃"
    return "未知"


def as_date(ts):
    if not ts:
        return ""
    return datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d")


def iso_to_date(value):
    if not value:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except ValueError:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def excerpt(post, limit=420):
    text = " ".join(((post.get("selftext") or "") or (post.get("title") or "")).split())
    return text[:limit]


def normalize_url(url):
    if not url:
        return ""
    url = str(url).strip()
    if url.startswith("/"):
        url = "https://www.reddit.com" + url
    return url.rstrip("/")


def load_rules(path):
    if not path or not Path(path).exists():
        return {}
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return data.get("subreddits", {})


def load_community_metadata(path, embedded):
    metadata = {}
    if embedded:
        metadata.update(embedded)
    if path and Path(path).exists():
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        if "community_metadata" in data:
            data = data.get("community_metadata") or {}
        elif "subreddits" in data:
            data = data.get("subreddits") or {}
        metadata.update(data)
    return metadata


def summarize_rules(sub, rules_data):
    item = rules_data.get(sub) or {}
    rules = item.get("rules") or []
    if not rules:
        if item.get("error"):
            return "needs manual verification; rules fetch failed"
        return "needs manual verification; no rules collected"
    names = [r.get("short_name") for r in rules if r.get("short_name")]
    summary = "; ".join(names[:6])
    if len(names) > 6:
        summary += f"; +{len(names)-6} more"
    return summary or "rules collected; review before posting"


def reply_risk(sub, pain_labels, intent_labels, rules_summary):
    high_risk_subs = {"piercing", "PiercingAdvice", "SkincareAddiction"}
    if sub in high_risk_subs:
        return "高"
    if "needs manual verification" in rules_summary:
        return "中"
    if any(x in joined(pain_labels) for x in ["敏感", "刺激", "发炎"]):
        return "中"
    if "意图：展示/晒图" in intent_labels and not any(x.startswith("痛点：") for x in pain_labels):
        return "低"
    return "低"


def product_design_labels(text):
    return match_labels(text, PRODUCT_DESIGN_PATTERNS)


def jewelry_type_labels(text):
    return match_labels(text, JEWELRY_TYPE_PATTERNS)


def is_jewelry_related(post, text, relevance):
    if relevance >= 3:
        return True
    if post.get("subreddit") in CORE_JEWELRY_SUBS:
        return True
    return bool(re.search(JEWELRY_TERMS, text, re.I))


def needs_product_design_review(design_labels, intent_labels, text):
    if not design_labels:
        return False
    if any(label in intent_labels for label in ["意图：抱怨/负面体验", "意图：求建议/求解决方案", "意图：购买研究/品牌比较"]):
        return True
    return bool(re.search(r"hurt|pain|heavy|allergic|sensitive|uncomfortable|irritation|swollen|itch|green skin", text, re.I))


def product_design_note(design_labels, product_labels):
    if not design_labels:
        return ""
    parts = []
    if "重量" in design_labels:
        parts.append("真实重量/拉扯反馈：请产品组判断耳饰重量、重心、耳垂拉扯和长时间佩戴负担。")
    if "过敏/敏感" in design_labels:
        parts.append("真实敏感/过敏反馈：请产品组判断材质、镀层、镍敏感、皮肤刺激和低敏表达边界。")
    if "佩戴舒适度" in design_labels:
        parts.append("真实舒适度反馈：请产品组判断耳针/耳堵/背扣、睡觉佩戴、松紧和摩擦痛点。")
    if "产品相关：Core-Lock/快拆结构" in product_labels:
        parts.append("同时验证快拆结构是否会增加松动、夹痛、丢件或清洁担忧。")
    return " ".join(parts)


def product_reply_decision(design_labels, risk):
    if not design_labels:
        return ""
    if risk == "高":
        return "先产品组判断事实，再由运营/创始人决定是否公开回复；避免医疗建议。"
    return "产品组提供产品事实/追问点，运营或创始人用研究式语气回复。"


def suggested_action(priority, reply_opp, risk, product_labels):
    if priority == "A" and reply_opp == "是" and risk == "低":
        return "温和回复"
    if priority == "A" and risk != "低":
        return "人工复核"
    if "产品相关：Core-Lock/快拆结构" in product_labels or "产品相关：旅行/胶囊衣橱" in product_labels:
        return "概念测试"
    if priority in {"A", "B"}:
        return "观察/收藏"
    return "归档"


def suggested_reply(post, pain_labels, intent_labels, product_labels, risk, rules_summary):
    if risk == "高":
        return "高风险帖：不要给医疗建议或硬广。建议先人工复核版规，只做中性澄清问题，或暂时观察。"
    prefix = "规则提醒：发帖前先复核 subreddit 规则；避免推广口吻，如涉及产品研究需透明说明身份。建议回复："
    if "产品相关：钛合金/低敏材质" in product_labels:
        return prefix + "我在做首饰材质方向的设计研究，想了解一个 tradeoff：遇到刺激或变色时，你更在意底层材质、镀层，还是品牌对材质说明是否透明？"
    if "产品相关：Core-Lock/快拆结构" in product_labels:
        return prefix + "这个问题很有意思，因为“容易更换”和“足够稳不容易掉”好像是相反需求。如果一个首饰结构可拆卸，什么细节会让你信任它？"
    if "产品相关：旅行/胶囊衣橱" in product_labels:
        return prefix + "好奇你怎么看旅行/胶囊衣橱里的配饰：你更倾向重复戴一件可靠单品，还是会使用可替换的小组件来减少携带数量？"
    if "痛点：价格/质量/信任" in pain_labels:
        return prefix + "从购买者角度看，什么证据会让较高价格变得合理：材质细节、耐久测试、可维修性、真实评价，还是保值能力？"
    if "意图：求建议/求解决方案" in intent_labels:
        return prefix + "这个问题很有参考价值。你这里最在意的是舒适度、材质、耐用性、百搭性，还是价格？"
    return ""


def problem_type(pain_labels, product_labels, design_labels, jewelry_types):
    labels = []
    priority_map = [
        ("过敏/敏感", ["过敏/敏感"], "痛点：敏感/刺激/发炎"),
        ("重量/拉扯", ["重量"], "痛点：舒适度/重量"),
        ("佩戴舒适度", ["佩戴舒适度"], "痛点：舒适度/重量"),
        ("掉色/氧化", [], "痛点：掉色/氧化/镀层磨损"),
        ("结构/扣件/丢件", [], "痛点：损坏/维修/拆装困难"),
        ("职场/场景切换", [], "痛点：造型/职场场景"),
        ("旅行/收纳", [], "痛点：旅行/收纳/缠绕"),
        ("价格/质量信任", [], "痛点：价格/质量/信任"),
    ]
    for label, design_needles, pain_needle in priority_map:
        if any(x in design_labels for x in design_needles) or pain_needle in pain_labels:
            labels.append(label)
    if "机会：模块化/可转换" in pain_labels or any("Core-Lock" in x for x in product_labels):
        labels.append("模块化机会")
    if not labels and jewelry_types:
        labels.append(jewelry_types[0])
    if not labels:
        labels.append("非核心/仅观察")
    return " / ".join(dict.fromkeys(labels[:3]))


def short_action(row):
    risk = row.get("Reply Risk")
    reply_opp = row.get("Reply Opportunity")
    source = row.get("Source Type")
    priority = row.get("Priority Tier")
    if risk == "高":
        return "人工复核"
    if reply_opp not in {"是", "可能"}:
        return "观察"
    if source == "rising" and priority in {"A", "B"}:
        return "优先回复"
    if priority == "A":
        return "准备回复"
    if priority == "B":
        return "观察后回复"
    return "收藏观察"


def suggested_reply_location(row):
    if row.get("Reply Risk") == "高" or row.get("Reply Opportunity") == "否":
        return "暂不回复"
    comments = row.get("Comment Count") or 0
    age_hours = row.get("帖子年龄（小时）") or 0
    if comments >= 80 and age_hours >= 24:
        return "评论回复"
    if comments >= 30 and row.get("Priority Tier") == "A":
        return "both"
    return "帖子回复"


def assignment_priority(row):
    risk = row.get("Reply Risk")
    if risk == "高":
        return "P2 待复核"
    if row.get("Source Type") == "rising" and row.get("Priority Tier") == "A":
        return "P0 今天回复"
    if row.get("Priority Tier") == "A":
        return "P1 两天内回复"
    if row.get("Priority Tier") == "B":
        return "P2 可回复"
    return "P3 观察"


def assignment_note(row):
    parts = []
    if row.get("Reply Risk") == "高":
        parts.append("高风险，先复核规则和语境。")
    if row.get("Source Type"):
        parts.append(f"来源：{row.get('Source Type')}")
    if row.get("Score Band"):
        parts.append(f"分数区间：{row.get('Score Band')}")
    if row.get("首饰类型"):
        parts.append(f"首饰类型：{row.get('首饰类型')}")
    if row.get("Rule-Aware Suggested Reply") and row.get("Reply Risk") != "高":
        parts.append(row.get("Rule-Aware Suggested Reply"))
    return " ".join(parts)[:600]


def assignment_reason(row):
    reasons = []
    source = row.get("Source Type")
    if source == "new":
        reasons.append("新帖，适合早期自然参与")
    elif source == "rising":
        reasons.append("Rising，讨论正在上升")
    elif source:
        reasons.append(f"来源：{source}")

    size = row.get("社区规模档位")
    activity = row.get("社区活跃度档位")
    if size in {"5万-10万", "10万-50万"}:
        reasons.append(f"中型垂直社区（{size}）")
    elif size:
        reasons.append(f"社区规模：{size}")
    if activity in {"高活跃", "中活跃"}:
        reasons.append(activity)

    comments = row.get("Comment Count") or 0
    comment_speed = row.get("每小时评论数") or 0
    age = row.get("帖子年龄（小时）") or 0
    if comments:
        reasons.append(f"{comments}条评论")
    if comment_speed:
        reasons.append(f"{comment_speed}/h评论速度")
    if age:
        reasons.append(f"{age}h内发布")

    if row.get("问题类型"):
        reasons.append(row.get("问题类型"))
    if row.get("Reply Risk") == "高":
        reasons.append("高风险，先复核")
    return "；".join(reasons[:6])


def community_rows(rows):
    by_sub = {}
    for row in rows:
        sub = row.get("Subreddit")
        if not sub or sub in by_sub:
            continue
        by_sub[sub] = row
    result = []
    for row in by_sub.values():
        result.append({key: row.get(key, "") for _, key in COMMUNITY_COLUMNS})
    result.sort(key=lambda r: (
        {"5万-10万": 0, "10万-50万": 1, "1万-5万": 2, "50万-100万": 3, "100万+": 4}.get(r.get("社区规模档位"), 9),
        r.get("Subreddit") or "",
    ))
    return result


def emotional_intensity(text, pain_labels, score, comments):
    high_words = r"hate|cry|cried|nightmare|disaster|regret|scam|desperate|impossible|worried|struggling|help|pain|hurts"
    val = 1
    if pain_labels:
        val += 1
    if re.search(high_words, text, re.I):
        val += 2
    if comments >= 50:
        val += 1
    if score >= 300:
        val += 1
    return min(5, val)


def jewelry_relevance(post, text):
    sub = post.get("subreddit")
    relevance = 0
    if sub in CORE_JEWELRY_SUBS:
        relevance += 3
    matches = re.findall(JEWELRY_TERMS, text, re.I)
    if matches:
        relevance += min(2, len(set(x.lower() for x in matches)))
    if post.get("query") in {"tarnish", "green skin", "sensitive ears", "titanium earrings", "office jewelry", "travel jewelry", "modular jewelry", "convertible jewelry", "versatile jewelry"}:
        relevance += 1
    return min(5, relevance)


def score_post(post, pain_labels, intent_labels, product_labels, intensity, relevance):
    score = 0
    strong_pain = [
        label for label in pain_labels
        if label not in {"痛点：造型/职场场景", "痛点：价格/质量/信任"}
    ]
    score += relevance * 6
    if pain_labels:
        score += 8 + min(9, len(pain_labels) * 2)
    if strong_pain:
        score += 6
    if "意图：抱怨/负面体验" in intent_labels:
        score += 12
    if "意图：求建议/求解决方案" in intent_labels:
        score += 9
    if "意图：购买研究/品牌比较" in intent_labels:
        score += 7
    score += min(8, len(product_labels) * 2)
    score += min(6, (post.get("num_comments") or 0) // 25)
    score += min(3, (post.get("score") or 0) // 350)
    score += intensity * 2
    if not pain_labels and "意图：展示/晒图" in intent_labels:
        score -= 15
    has_actionable_intent = any(
        label in intent_labels
        for label in ["意图：抱怨/负面体验", "意图：求建议/求解决方案", "意图：购买研究/品牌比较", "意图：保养/维修/护理"]
    )
    has_strong_product_signal = any(
        label in product_labels
        for label in ["产品相关：钛合金/低敏材质", "产品相关：Core-Lock/快拆结构", "产品相关：职场到社交切换", "产品相关：旅行/胶囊衣橱"]
    )
    if not strong_pain and not has_actionable_intent and not has_strong_product_signal:
        score = min(score, 38)
    elif not strong_pain and "意图：展示/晒图" in intent_labels and not has_strong_product_signal:
        score = min(score, 45)
    elif not strong_pain and not has_strong_product_signal:
        score = min(score, 68)
    if not strong_pain and pain_labels == ["痛点：造型/职场场景"] and "意图：求建议/求解决方案" not in intent_labels:
        score = min(score, 45)
    if relevance <= 1:
        score = min(score, 24)
    elif relevance == 2:
        score = min(score, 45)
    elif relevance == 3:
        score = min(score, 68)
    return max(0, min(100, score))


def tier(score):
    if score >= 70:
        return "A"
    if score >= 50:
        return "B"
    if score >= 30:
        return "C"
    return "D"


def score_band(score):
    return tier(score)


def hours_since(ts, reference_dt):
    if not ts:
        return None
    created = datetime.fromtimestamp(ts, timezone.utc)
    hours = (reference_dt - created).total_seconds() / 3600
    return max(hours, 0.1)


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def compact_worksheet(ws, row_height=22, freeze_panes=None, zoom=90, wrap_data=False):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    vertical = "top" if wrap_data else "center"
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = row_height
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical=vertical, wrap_text=wrap_data)


def apply_width_overrides(ws, overrides):
    for col, width in overrides.items():
        ws.column_dimensions[col].width = width


def hide_columns_by_header(ws, labels):
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    for label in labels:
        col_idx = header_to_col.get(label)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def build_rows(posts, rules_data, collection_date, reference_dt, community_metadata=None):
    community_metadata = community_metadata or {}
    rows = []
    for post in posts:
        text = f"{post.get('title') or ''} {post.get('selftext') or ''} {post.get('query') or ''}"
        pain = match_labels(text, TAG_PATTERNS)
        intent = match_labels(text, INTENT_PATTERNS)
        product = match_labels(text, PRODUCT_PATTERNS)
        design_labels = product_design_labels(text)
        relevance = jewelry_relevance(post, text)
        type_labels = jewelry_type_labels(text)
        jewelry_related = is_jewelry_related(post, text, relevance)
        sub = post.get("subreddit")
        sub_meta = community_metadata.get(sub) or community_metadata.get((sub or "").lower()) or {}
        api_subscribers = sub_meta.get("api_subscribers") or sub_meta.get("subscribers")
        sidebar_subscribers = sub_meta.get("sidebar_subscribers")
        subscribers = sidebar_subscribers or api_subscribers
        active_users = sub_meta.get("weekly_active_users") or sub_meta.get("active_user_count") or sub_meta.get("accounts_active")
        weekly_contributions = sub_meta.get("weekly_contributions")
        age_hours = hours_since(post.get("created_utc"), reference_dt)
        rules_summary = summarize_rules(sub, rules_data)
        intensity = emotional_intensity(text, pain, post.get("score") or 0, post.get("num_comments") or 0)
        risk = reply_risk(sub, pain, intent, rules_summary)
        reply_opp = "是" if pain and ("意图：求建议/求解决方案" in intent or "意图：抱怨/负面体验" in intent or product) else "可能" if product else "否"
        importance = score_post(post, pain, intent, product, intensity, relevance)
        priority = tier(importance)
        action = suggested_action(priority, reply_opp, risk, product)
        design_review = jewelry_related and needs_product_design_review(design_labels, intent, text)
        opportunity = []
        if "产品相关：钛合金/低敏材质" in product:
            opportunity.append("材质信任")
        if "产品相关：Core-Lock/快拆结构" in product:
            opportunity.append("结构信任")
        if "产品相关：职场到社交切换" in product:
            opportunity.append("场景切换")
        if "产品相关：旅行/胶囊衣橱" in product:
            opportunity.append("旅行/少而精")
        if "产品相关：质量/价值证明" in product:
            opportunity.append("价值证明")
        suiji = []
        if "产品相关：钛合金/低敏材质" in product:
            suiji.append("使用谨慎的 Grade 5 钛合金/低敏材质表达，避免绝对医疗承诺。")
        if "产品相关：Core-Lock/快拆结构" in product:
            suiji.append("验证 Core-Lock 信任点：好换、锁得稳、不易丢件。")
        if "产品相关：旅行/胶囊衣橱" in product:
            suiji.append("测试“更少首饰覆盖更多场景”的旅行/胶囊衣橱叙事。")
        if "产品相关：职场到社交切换" in product:
            suiji.append("将产品表达为不同场景中的自我表达切换。")
        if "产品相关：质量/价值证明" in product:
            suiji.append("通过材质、结构、耐久测试、可维修性解释价格。")
        row = {
            "Reddit Post ID": post.get("id") or post.get("name") or "",
            "Importance Score": importance,
            "Score Band": score_band(importance),
            "Priority Tier": priority,
            "问题类型": problem_type(pain, product, design_labels, type_labels),
            "Subreddit": sub,
            "Subreddit Group": SUBREDDIT_GROUPS.get(sub, "Other"),
            "API订阅人数": api_subscribers if api_subscribers not in (None, "") else "",
            "Sidebar订阅人数": sidebar_subscribers if sidebar_subscribers not in (None, "") else "",
            "社区成员数": subscribers if subscribers not in (None, "") else "",
            "社区成员数显示": compact_count(subscribers),
            "页面周活跃用户数": active_users if active_users not in (None, "") else "",
            "页面周活跃用户数显示": compact_count(active_users),
            "页面周贡献数": weekly_contributions if weekly_contributions not in (None, "") else "",
            "页面周贡献数显示": compact_count(weekly_contributions),
            "社区当前活跃人数": active_users if active_users not in (None, "") else "",
            "社区当前活跃人数显示": compact_count(active_users),
            "社区活跃比例": active_ratio(active_users, subscribers),
            "社区规模档位": size_band(subscribers),
            "社区活跃度档位": activity_band(active_users, subscribers),
            "页面成员标签": sub_meta.get("page_subscribers_text") or "",
            "页面活跃标签": sub_meta.get("page_currently_viewing_text") or "",
            "社区数据来源": sub_meta.get("source") or sub_meta.get("page_source") or "",
            "社区数据状态": sub_meta.get("status") or "未知/待人工补充",
            "Post Title": post.get("title"),
            "Post URL": normalize_url(post.get("permalink")),
            "Author": post.get("author"),
            "Collection Date": collection_date,
            "Post Published Date": as_date(post.get("created_utc")),
            "Reddit Score": post.get("score"),
            "Comment Count": post.get("num_comments"),
            "帖子年龄（小时）": round(age_hours, 1) if age_hours is not None else "",
            "每小时得分": round((post.get("score") or 0) / age_hours, 2) if age_hours else "",
            "每小时评论数": round((post.get("num_comments") or 0) / age_hours, 2) if age_hours else "",
            "Jewelry Relevance": relevance,
            "Source Type": post.get("source_kind"),
            "Query": post.get("query") or "",
            "是否首饰类": "是" if jewelry_related else "否",
            "首饰类型": joined(type_labels),
            "Pain Clusters": joined(pain),
            "Intent Tags": joined(intent),
            "Product Fit Tags": joined(product),
            "产品组痛点类型": joined(design_labels),
            "是否产品组处理": "是" if design_review else "否",
            "负责人": "产品组" if design_review else "运营/创始人",
            "产品组备注": product_design_note(design_labels, product) if design_review else "",
            "回复归属建议": product_reply_decision(design_labels, risk) if design_review else "",
            "Opportunity Type": joined(opportunity),
            "Emotional Intensity": intensity,
            "Reply Opportunity": reply_opp,
            "Reply Risk": risk,
            "Seed User Fit": "高" if priority == "A" and reply_opp in {"是", "可能"} else "中" if priority == "B" else "低",
            "Recommended Action": action,
            "SUIJI Implication": " ".join(suiji),
            "Subreddit Rules Summary": rules_summary,
            "Rule-Aware Suggested Reply": suggested_reply(post, pain, intent, product, risk, rules_summary),
            "Excerpt": excerpt(post),
        }
        row.update(action_fields(row))
        rows.append(row)
    rows.sort(key=lambda r: (r["Importance Score"], r["Comment Count"] or 0, r["Reddit Score"] or 0), reverse=True)
    for i, row in enumerate(rows, start=1):
        row["Rank"] = i
    return rows


def append_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_header(ws)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


POST_COLUMNS = [
    ("排名", "Rank"),
    ("Reddit Post ID", "Reddit Post ID"),
    ("重要性分数", "Importance Score"),
    ("分数区间", "Score Band"),
    ("优先级", "Priority Tier"),
    ("问题类型", "问题类型"),
    ("处理优先级", "处理优先级"),
    ("回复顺序建议", "回复顺序建议"),
    ("优先理由", "优先理由"),
    ("Subreddit", "Subreddit"),
    ("Subreddit 分类", "Subreddit Group"),
    ("API订阅人数", "API订阅人数"),
    ("Sidebar订阅人数", "Sidebar订阅人数"),
    ("社区成员数", "社区成员数"),
    ("社区成员数显示", "社区成员数显示"),
    ("页面周活跃用户数", "页面周活跃用户数"),
    ("页面周活跃用户数显示", "页面周活跃用户数显示"),
    ("页面周贡献数", "页面周贡献数"),
    ("页面周贡献数显示", "页面周贡献数显示"),
    ("社区当前活跃人数", "社区当前活跃人数"),
    ("社区当前活跃人数显示", "社区当前活跃人数显示"),
    ("社区活跃比例", "社区活跃比例"),
    ("社区规模档位", "社区规模档位"),
    ("社区活跃度档位", "社区活跃度档位"),
    ("页面成员标签", "页面成员标签"),
    ("页面活跃标签", "页面活跃标签"),
    ("社区数据来源", "社区数据来源"),
    ("社区数据状态", "社区数据状态"),
    ("帖子标题", "Post Title"),
    ("帖子链接", "Post URL"),
    ("作者", "Author"),
    ("收集日期", "Collection Date"),
    ("帖子发布日期", "Post Published Date"),
    ("Reddit 原帖分数", "Reddit Score"),
    ("评论数", "Comment Count"),
    ("帖子年龄（小时）", "帖子年龄（小时）"),
    ("每小时得分", "每小时得分"),
    ("每小时评论数", "每小时评论数"),
    ("首饰相关性", "Jewelry Relevance"),
    ("来源类型", "Source Type"),
    ("搜索关键词", "Query"),
    ("是否首饰类", "是否首饰类"),
    ("首饰类型", "首饰类型"),
    ("痛点标签", "Pain Clusters"),
    ("用户意图标签", "Intent Tags"),
    ("产品相关标签", "Product Fit Tags"),
    ("产品组痛点类型", "产品组痛点类型"),
    ("是否产品组处理", "是否产品组处理"),
    ("负责人", "负责人"),
    ("产品组备注", "产品组备注"),
    ("回复归属建议", "回复归属建议"),
    ("机会类型", "Opportunity Type"),
    ("情绪强度", "Emotional Intensity"),
    ("回复机会", "Reply Opportunity"),
    ("回复风险", "Reply Risk"),
    ("种子用户匹配度", "Seed User Fit"),
    ("建议动作", "Recommended Action"),
    ("SUIJI 启发", "SUIJI Implication"),
    ("Subreddit 规则摘要", "Subreddit Rules Summary"),
    ("参考规则和帖子内容的回复建议", "Rule-Aware Suggested Reply"),
    ("帖子摘要", "Excerpt"),
]

ACTION_COLUMNS = [
    ("处理优先级", "处理优先级"),
    ("回复顺序建议", "回复顺序建议"),
    ("优先理由", "优先理由"),
    ("建议动作", "Recommended Action"),
    ("负责人", "负责人"),
    ("回复风险", "Reply Risk"),
    ("回复机会", "Reply Opportunity"),
    ("分数区间", "Score Band"),
    ("重要性分数", "Importance Score"),
    ("来源类型", "Source Type"),
    ("Subreddit", "Subreddit"),
    ("社区规模档位", "社区规模档位"),
    ("社区活跃度档位", "社区活跃度档位"),
    ("帖子标题", "Post Title"),
    ("帖子链接", "Post URL"),
    ("帖子发布日期", "Post Published Date"),
    ("帖子年龄（小时）", "帖子年龄（小时）"),
    ("每小时评论数", "每小时评论数"),
    ("每小时得分", "每小时得分"),
    ("是否首饰类", "是否首饰类"),
    ("首饰类型", "首饰类型"),
    ("痛点标签", "Pain Clusters"),
    ("用户意图标签", "Intent Tags"),
    ("产品相关标签", "Product Fit Tags"),
    ("产品组痛点类型", "产品组痛点类型"),
    ("产品组备注", "产品组备注"),
    ("SUIJI 启发", "SUIJI Implication"),
    ("参考规则和帖子内容的回复建议", "Rule-Aware Suggested Reply"),
    ("帖子摘要", "Excerpt"),
]

PRODUCT_DESIGN_COLUMNS = [
    ("处理优先级", "处理优先级"),
    ("优先理由", "优先理由"),
    ("Subreddit", "Subreddit"),
    ("社区规模档位", "社区规模档位"),
    ("社区活跃度档位", "社区活跃度档位"),
    ("帖子标题", "Post Title"),
    ("帖子链接", "Post URL"),
    ("帖子发布日期", "Post Published Date"),
    ("是否首饰类", "是否首饰类"),
    ("首饰类型", "首饰类型"),
    ("产品组痛点类型", "产品组痛点类型"),
    ("痛点标签", "Pain Clusters"),
    ("用户意图标签", "Intent Tags"),
    ("产品相关标签", "Product Fit Tags"),
    ("负责人", "负责人"),
    ("产品组备注", "产品组备注"),
    ("回复归属建议", "回复归属建议"),
    ("SUIJI 启发", "SUIJI Implication"),
    ("参考规则和帖子内容的回复建议", "Rule-Aware Suggested Reply"),
    ("帖子摘要", "Excerpt"),
]

DAILY_RISING_COLUMNS = ACTION_COLUMNS + [("是否建议今天处理", "是否建议今天处理")]

REPLY_ASSIGNMENT_COLUMNS = [
    ("优先级", "回复分配优先级"),
    ("来源", "Source Type"),
    ("社区", "Subreddit"),
    ("社区成员数", "社区成员数显示"),
    ("社区规模档位", "社区规模档位"),
    ("社区活跃度", "社区活跃度档位"),
    ("Reddit原帖分数", "Reddit Score"),
    ("评论数", "Comment Count"),
    ("帖子年龄h", "帖子年龄（小时）"),
    ("评论速度/h", "每小时评论数"),
    ("问题类型", "问题类型"),
    ("首饰类型", "首饰类型"),
    ("帖子标题", "Post Title"),
    ("原帖链接", "Post URL"),
    ("建议回复位置", "建议回复位置"),
    ("建议动作", "回复分配建议动作"),
    ("为什么值得回", "为什么值得回"),
    ("负责人名", "负责人名"),
    ("是否回复", "是否回复"),
    ("实际回复类型", "实际回复类型"),
    ("回复内容", "回复内容"),
]

PROCESSED_COLUMNS = [
    ("帖子标题", "帖子标题"),
    ("原帖链接", "原帖链接"),
    ("负责人名", "负责人名"),
    ("是否回复", "是否回复"),
    ("实际回复类型", "实际回复类型"),
    ("回复内容", "回复内容"),
    ("回复日期", "回复日期"),
    ("跟进状态", "跟进状态"),
    ("备注", "人工备注"),
]

COMMUNITY_COLUMNS = [
    ("Subreddit", "Subreddit"),
    ("社区分类", "Subreddit Group"),
    ("成员数", "社区成员数"),
    ("成员数显示", "社区成员数显示"),
    ("规模档位", "社区规模档位"),
    ("周活跃用户数", "页面周活跃用户数"),
    ("周活跃显示", "页面周活跃用户数显示"),
    ("周贡献数", "页面周贡献数"),
    ("周贡献显示", "页面周贡献数显示"),
    ("活跃度档位", "社区活跃度档位"),
    ("社区数据状态", "社区数据状态"),
    ("规则摘要", "Subreddit Rules Summary"),
]

ACTION_PRIORITY_ORDER = {
    "P0 今天优先回复": 0,
    "P1 产品组先看": 1,
    "P2 本周准备回复": 2,
    "P3 收藏观察": 3,
    "P4 暂不公开回复": 4,
    "P0 今天回复": 0,
    "P1 两天内回复": 1,
    "P2 可回复": 2,
    "P2 待复核": 3,
    "P3 观察": 4,
}

COLUMN_GUIDE = {
    "处理优先级": ("系统给出的行动优先级，P0 最急，P4 不建议公开回复。", "核心", "运营/创始人", "先筛 P0/P1/P2；P4 默认不回复。"),
    "优先级": ("回复分配表中的处理优先级，P0 最急，P3 观察。", "核心", "所有人", "先处理 P0/P1，P2 人工判断，P3 观察。"),
    "问题类型": ("这条帖子主要对应的首饰问题或机会类型。", "核心", "所有人", "按过敏、重量、舒适度、结构、职场、旅行等类型分配。"),
    "建议回复位置": ("系统建议回复原帖、回复评论、两者都做，或暂不回复。", "核心", "所有人", "发布前人工确认；评论回复需要实际打开帖子找合适评论。"),
    "实际回复类型": ("负责人实际采取的回复方式。", "核心", "所有人", "填写未回复、帖子回复、评论回复、both、不回复。"),
    "负责人名": ("被分配负责这条互动的人。", "核心", "所有人", "由团队手动填写。"),
    "是否回复": ("这条候选是否已经处理。", "核心", "所有人", "填写否、是、不适合回复、待复核。"),
    "回复内容": ("实际发出或准备发出的回复内容。", "核心", "所有人", "回写总表，便于避免重复和复盘。"),
    "备注": ("人工备注或系统提醒。", "高", "所有人", "写规则提醒、语境判断、暂缓原因或后续动作。"),
    "原帖链接": ("Reddit 原帖链接，也是回写总表的匹配键。", "核心", "所有人", "所有回复前都要打开原帖核对。"),
    "回复顺序建议": ("说明 Rising、New、产品组和人工复核之间的处理顺序。", "核心", "运营/创始人", "Daily 先看 Rising，再看高度相关 New。"),
    "优先理由": ("解释为什么该帖被放进当前优先级。", "核心", "所有人", "用于快速理解系统判断，人工可覆盖。"),
    "排名": ("按重要性分数、评论数、Reddit 原帖分数综合排序后的序号。", "中", "研究/分析", "辅助定位，不要只看排名做回复。"),
    "重要性分数": ("SUIJI 调研价值分，0-100，区别于 Reddit 热度。", "核心", "所有人", "A/B 优先看；低分通常归档。"),
    "分数区间": ("A=70+，B=50-69，C=30-49，D<30。", "核心", "所有人", "先筛 A/B，C/D 只作背景库。"),
    "优先级": ("与分数区间一致，用于旧版兼容。", "中", "研究/分析", "可与分数区间一起筛。"),
    "Subreddit": ("帖子所在社区。", "高", "所有人", "结合社区规则和社区体量判断回复方式。"),
    "Subreddit 分类": ("社区的业务分类，如首饰主社区、穿搭/职场、旅行轻量化。", "中", "研究/分析", "用来分组看痛点来源。"),
    "API订阅人数": ("来自 Reddit about.json 的 subscribers。", "辅助", "研究/分析", "只作社区体量参考。"),
    "Sidebar订阅人数": ("来自页面 Subscriber Info 的订阅人数。", "辅助", "研究/分析", "和 API 订阅人数一起看，可能不完全一致。"),
    "社区成员数": ("当前可用的最佳社区成员数。", "中", "所有人", "判断社区体量，不直接决定是否回复。"),
    "社区成员数显示": ("成员数的 K/M 简写。", "辅助", "所有人", "用于阅读友好。"),
    "页面周活跃用户数": ("页面 header 抓到的 weekly-active-users。", "中", "研究/分析", "判断社区近期活跃度。"),
    "页面周活跃用户数显示": ("页面周活跃用户数的 K/M 简写。", "辅助", "所有人", "用于快速浏览。"),
    "页面周贡献数": ("页面 header 抓到的 weekly-contributions。", "中", "研究/分析", "衡量近期贡献/发言规模。"),
    "页面周贡献数显示": ("页面周贡献数的 K/M 简写。", "辅助", "所有人", "用于快速浏览。"),
    "社区当前活跃人数": ("当前版本沿用页面周活跃用户数作为活跃上下文。", "辅助", "研究/分析", "不要当作实时在线人数。"),
    "社区当前活跃人数显示": ("当前活跃人数的 K/M 简写。", "辅助", "所有人", "用于快速浏览。"),
    "社区活跃比例": ("活跃人数除以成员数的近似比例。", "中", "研究/分析", "辅助判断小社区是否很活跃。"),
    "社区规模档位": ("社区体量分层，如 50万-100万、10万-50万。", "高", "所有人", "判断一个痛点出现在哪类社区。"),
    "社区活跃度档位": ("高/中/低活跃的粗分层。", "高", "所有人", "高活跃社区的回复要更谨慎。"),
    "页面成员标签": ("页面显示的成员称呼，如 Piercing enthusiasts。", "辅助", "研究/分析", "用于理解社区语言。"),
    "页面活跃标签": ("页面显示的活跃称呼，如 Enthusiasts contributing。", "辅助", "研究/分析", "用于理解社区语言。"),
    "社区数据来源": ("社区元数据来自 API、页面，或混合来源。", "辅助", "研究/分析", "排查数据可信度。"),
    "社区数据状态": ("社区元数据抓取是否成功或待人工补充。", "中", "研究/分析", "失败时不要过度解读社区体量。"),
    "帖子标题": ("Reddit 帖子标题。", "核心", "所有人", "快速判断是否需要打开原帖。"),
    "帖子链接": ("Reddit 原帖链接。", "核心", "所有人", "所有引用和回复前都要打开核对。"),
    "作者": ("Reddit 作者名。", "辅助", "运营/创始人", "用于种子用户跟踪，不建议私信骚扰。"),
    "收集日期": ("本工作流抓到该帖的日期。", "中", "研究/分析", "用于追踪周报/日报来源。"),
    "帖子发布日期": ("Reddit 原帖发布时间。", "高", "所有人", "判断是否还适合回复。"),
    "Reddit 原帖分数": ("Reddit 原生热度分，不等于 SUIJI 重要性。", "中", "所有人", "只作辅助，不要按它单独排序。"),
    "评论数": ("帖子评论数量。", "中", "所有人", "评论多说明讨论充分，但仍需看相关性。"),
    "帖子年龄（小时）": ("帖子从发布到抓取时经过的小时数。", "高", "运营/创始人", "越新越适合日报监控。"),
    "每小时得分": ("Reddit 原帖分数除以帖子年龄。", "高", "运营/创始人", "判断热度速度，配合 Rising 使用。"),
    "每小时评论数": ("评论数除以帖子年龄。", "高", "运营/创始人", "判断讨论速度，日报优先看高值。"),
    "首饰相关性": ("1-5 的首饰相关性评分。", "核心", "所有人", "低于 3 通常不要回复。"),
    "来源类型": ("帖子来自 new、rising、top_week 或 search。", "核心", "运营/创始人", "Rising 优先于普通 New。"),
    "搜索关键词": ("如果来自 search，记录触发关键词。", "中", "研究/分析", "看哪些关键词带来有效帖子。"),
    "是否首饰类": ("帖子是否被判定为首饰相关。", "核心", "产品组", "产品组只看“是”。"),
    "首饰类型": ("耳饰、戒指、项链、穿孔饰品、材质/过敏等分类。", "核心", "产品组", "按品类分给对应设计/产品同事。"),
    "痛点标签": ("掉色、敏感、重量、损坏、职场、旅行等痛点聚类。", "核心", "所有人", "报告和产品分诊的核心筛选字段。"),
    "用户意图标签": ("抱怨、求建议、购买研究、保养维修、展示等。", "核心", "运营/创始人", "求建议/抱怨更适合回复。"),
    "产品相关标签": ("与钛合金、Core-Lock、场景切换、旅行等 SUIJI 方向的关系。", "高", "研究/分析", "用于判断产品机会。"),
    "产品组痛点类型": ("重量、过敏/敏感、佩戴舒适度三类产品组重点。", "核心", "产品组", "命中后先给产品组判断。"),
    "是否产品组处理": ("是否需要产品组先看。", "核心", "产品组", "筛“是”生成产品组表。"),
    "负责人": ("建议由产品组、运营或创始人负责下一步。", "核心", "所有人", "用来分工。"),
    "产品组备注": ("产品组需要判断的事实、风险或设计问题。", "核心", "产品组", "产品组日报/分诊表重点看。"),
    "回复归属建议": ("产品组提供事实，运营/创始人负责公开回复的建议。", "高", "所有人", "避免产品组直接硬回复。"),
    "机会类型": ("材质信任、结构信任、场景切换、旅行/少而精、价值证明等。", "中", "研究/分析", "用于周报机会章节。"),
    "情绪强度": ("1-5 的情绪强弱。", "中", "研究/分析", "高情绪痛点更值得关注。"),
    "回复机会": ("是否存在合适的公开回复机会。", "核心", "运营/创始人", "优先筛“是/可能”。"),
    "回复风险": ("低/中/高，考虑版规、医疗敏感、穿孔建议等风险。", "核心", "运营/创始人", "高风险不直接回复。"),
    "种子用户匹配度": ("该作者/帖子是否可能成为早期互动对象。", "中", "运营/创始人", "高匹配也不能私信骚扰。"),
    "建议动作": ("温和回复、人工复核、概念测试、观察/收藏、归档。", "核心", "所有人", "转化为下一步动作。"),
    "SUIJI 启发": ("这条帖子对 SUIJI 产品、材质、结构或叙事的启发。", "高", "研究/分析", "周报和产品会讨论用。"),
    "Subreddit 规则摘要": ("抓取到的社区规则摘要。", "高", "运营/创始人", "回复前必须复核。"),
    "参考规则和帖子内容的回复建议": ("结合帖子和规则生成的温和回复方向。", "核心", "运营/创始人", "只能作为草稿，发布前人工改写。"),
    "帖子摘要": ("标题/正文的短摘录。", "核心", "所有人", "快速理解帖子，不替代打开原帖。"),
    "是否建议今天处理": ("Daily Rising 专用，判断今天是否需要处理。", "核心", "运营/创始人", "筛“是”，再按 P0/P1 处理。"),
}


def append_post_table(ws, rows):
    labels = [label for label, _ in POST_COLUMNS]
    keys = [key for _, key in POST_COLUMNS]
    ws.append(labels)
    for row in rows:
        ws.append([row.get(key, "") for key in keys])
    style_header(ws)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_mapped_table(ws, columns, rows):
    labels = [label for label, _ in columns]
    keys = [key for _, key in columns]
    ws.append(labels)
    for row in rows:
        ws.append([row.get(key, "") for key in keys])
    style_header(ws)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def action_sort_key(row):
    return (
        ACTION_PRIORITY_ORDER.get(row.get("处理优先级"), 9),
        -(row.get("Importance Score") or 0),
        -(row.get("每小时评论数") or 0),
        -(row.get("每小时得分") or 0),
    )


def action_queue_rows(rows):
    candidates = [
        r for r in rows
        if r.get("处理优先级") in {"P0 今天优先回复", "P1 产品组先看", "P2 本周准备回复", "P3 收藏观察"}
    ]
    return sorted(candidates, key=action_sort_key)


def is_reply_assignment_candidate(row):
    relevance = row.get("Jewelry Relevance") or 0
    if relevance < 3:
        return False
    if row.get("Reply Opportunity") == "否" and row.get("Reply Risk") != "高":
        return False
    if row.get("Intent Tags") == "意图：展示/晒图" and not row.get("Pain Clusters"):
        return False
    has_real_signal = bool(row.get("Pain Clusters") or row.get("Product Fit Tags") or row.get("产品组痛点类型"))
    if not has_real_signal:
        return False
    if row.get("Reply Risk") == "高":
        return row.get("Score Band") in {"A", "B"} and relevance >= 4
    if row.get("Source Type") == "rising" and relevance >= 4:
        return True
    return row.get("Score Band") in {"A", "B"} and row.get("Reply Opportunity") in {"是", "可能"}


def reply_assignment_sort_key(row):
    return (
        ACTION_PRIORITY_ORDER.get(row.get("回复分配优先级"), 9),
        -(row.get("Importance Score") or 0),
        -(row.get("每小时评论数") or 0),
        -(row.get("每小时得分") or 0),
    )


def reply_assignment_rows(rows, processed_urls=None, limit=80):
    processed_urls = processed_urls or set()
    assignment_rows = []
    for row in rows:
        url = normalize_url(row.get("Post URL"))
        if url and url in processed_urls:
            continue
        if not is_reply_assignment_candidate(row):
            continue
        item = dict(row)
        item["回复分配优先级"] = assignment_priority(row)
        item["建议回复位置"] = suggested_reply_location(row)
        item["回复分配建议动作"] = short_action(row)
        item["为什么值得回"] = assignment_reason(row)
        item["负责人名"] = ""
        item["是否回复"] = "待复核" if row.get("Reply Risk") == "高" else "否"
        item["实际回复类型"] = "未回复"
        item["回复内容"] = ""
        item["回复分配备注"] = assignment_note(row)
        assignment_rows.append(item)
    assignment_rows.sort(key=reply_assignment_sort_key)
    return assignment_rows[:limit]


def find_header(headers, *candidates):
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate)
    return None


def load_processed_rows(master_path):
    if not master_path or not Path(master_path).exists():
        return [], set()
    wb = load_workbook(master_path, read_only=True, data_only=True)
    sheet = None
    for name in ["Master Posts", "全量数据库", "Posts"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(x).strip() if x is not None else "" for x in next(rows_iter)]
    except StopIteration:
        return [], set()
    url_idx = find_header(headers, "原帖链接", "帖子链接", "Post URL")
    title_idx = find_header(headers, "帖子标题", "Post Title")
    owner_idx = find_header(headers, "负责人名", "跟进负责人", "负责人")
    replied_idx = find_header(headers, "是否回复")
    reply_type_idx = find_header(headers, "实际回复类型")
    reply_idx = find_header(headers, "回复内容")
    date_idx = find_header(headers, "回复日期", "最后动作日期")
    status_idx = find_header(headers, "跟进状态")
    note_idx = find_header(headers, "人工备注", "备注")
    processed = []
    processed_urls = set()
    for values in rows_iter:
        get = lambda idx: values[idx] if idx is not None and idx < len(values) and values[idx] is not None else ""
        url = normalize_url(get(url_idx))
        replied = str(get(replied_idx)).strip()
        reply_type = str(get(reply_type_idx)).strip()
        status = str(get(status_idx)).strip()
        is_done = (
            replied in {"是", "不适合回复"}
            or reply_type in {"帖子回复", "评论回复", "both", "不回复"}
            or status in {"已回复", "不适合回复", "已归档"}
        )
        if not is_done:
            continue
        if url:
            processed_urls.add(url)
        processed.append({
            "帖子标题": get(title_idx),
            "原帖链接": url,
            "负责人名": get(owner_idx),
            "是否回复": replied,
            "实际回复类型": reply_type,
            "回复内容": get(reply_idx),
            "回复日期": get(date_idx),
            "跟进状态": status,
            "人工备注": get(note_idx),
        })
    return processed, processed_urls


def append_reply_assignment_field_guide(ws):
    rows = [
        ("优先级", "系统", "是", "P0今天回复/P1两天内回复/P2可回复/P2待复核/P3观察", "先处理 P0/P1；P2 人工判断；P3 只观察。", "P0 今天回复"),
        ("来源", "系统", "是", "new/rising/search/top_week", "new 适合早期参与；rising 表示正在升温；search/top_week 更多用于观察。", "new"),
        ("社区", "系统", "是", "Subreddit 名称", "筛选目标社区，回复前打开社区规则。", "jewelrymaking"),
        ("社区成员数", "系统", "否", "自动抓取", "判断社区体量；中型垂直社区常更适合早期互动。", "99K"),
        ("社区规模档位", "系统", "是", "1万以下/1万-5万/5万-10万/10万-50万/50万-100万/100万+", "可直接筛 5万-10万 或 10万-50万。", "5万-10万"),
        ("社区活跃度", "系统", "是", "高活跃/中活跃/低活跃/未知", "高/中活跃更值得投入，但高活跃社区回复也更需要谨慎。", "高活跃"),
        ("Reddit原帖分数", "系统", "否", "数字", "Reddit 原生 score，不等于我们的重要性；新帖低分不一定差。", "4"),
        ("评论数", "系统", "是", "数字", "判断是否已有讨论。评论多时可考虑评论回复。", "10"),
        ("帖子年龄h", "系统", "是", "小时", "越新越适合早期自然回复。", "3.5"),
        ("评论速度/h", "系统", "是", "数字", "判断讨论增长速度。", "0.58"),
        ("问题类型", "系统", "是", "过敏/重量/舒适/掉色/结构/职场/旅行/价格/模块化等", "用于判断回复角度。", "结构/扣件/丢件"),
        ("首饰类型", "系统", "否", "耳饰/戒指/项链/穿孔饰品等", "用于判断是否贴合 SUIJI 和回复语境。", "项链/吊坠"),
        ("帖子标题", "系统", "是", "Reddit 标题", "快速判断是否打开原帖。", "how do i fix this clasp-less necklace??"),
        ("原帖链接", "系统", "是", "URL", "回复前必须打开原帖；同步总表时用它匹配。", "https://www.reddit.com/..."),
        ("建议回复位置", "系统", "是", "帖子回复/评论回复/both/暂不回复", "评论多且讨论分散时优先找评论切入；高风险默认暂不回复。", "帖子回复"),
        ("建议动作", "系统", "是", "优先回复/准备回复/观察后回复/人工复核/观察", "给负责人下一步动作。", "优先回复"),
        ("为什么值得回", "系统", "是", "自动生成", "用一句话解释来源、社区、热度、问题类型等判断。", "new；中型垂直社区；10条评论；结构问题"),
        ("负责人名", "人工", "是", "自由填写", "分配谁去处理。", "Wendy"),
        ("是否回复", "人工", "是", "否/是/不适合回复/待复核", "处理后更新状态。", "是"),
        ("实际回复类型", "人工", "是", "未回复/帖子回复/评论回复/both/不回复", "记录实际回复位置。", "评论回复"),
        ("回复内容", "人工", "否", "自由填写", "记录实际发出或准备发出的回复内容，会同步回总表。", "Thanks for sharing..."),
        ("社区库", "系统", "否", "独立 sheet", "用于筛选和复核社区体量、活跃度、规则摘要。", "r/PiercingAdvice｜99K｜高活跃"),
        ("已处理", "系统/人工", "否", "独立 sheet", "来自总表的已回复/不回复记录，避免重复处理。", "已回复"),
    ]
    append_table(
        ws,
        ["字段名", "谁填写", "是否必填", "可选值", "怎么判断", "示例"],
        [{"字段名": a, "谁填写": b, "是否必填": c, "可选值": d, "怎么判断": e, "示例": f} for a, b, c, d, e, f in rows],
    )
    set_widths(ws, [20, 14, 12, 38, 70, 36])


def add_reply_assignment_validations(ws, max_rows):
    if max_rows < 2:
        max_rows = 200
    reply_status = DataValidation(type="list", formula1='"否,是,不适合回复,待复核"', allow_blank=True)
    reply_type = DataValidation(type="list", formula1='"未回复,帖子回复,评论回复,both,不回复"', allow_blank=True)
    ws.add_data_validation(reply_status)
    ws.add_data_validation(reply_type)
    reply_status.add(f"S2:S{max_rows}")
    reply_type.add(f"T2:T{max_rows}")


def build_reply_assignment_workbook(wb, rows, master_path, limit):
    processed_rows, processed_urls = load_processed_rows(master_path)
    assignment = reply_assignment_rows(rows, processed_urls=processed_urls, limit=limit)
    pending_ws = wb.active
    pending_ws.title = "待回复"
    append_mapped_table(pending_ws, REPLY_ASSIGNMENT_COLUMNS, assignment)
    set_widths(pending_ws, [16, 12, 18, 14, 14, 14, 14, 10, 12, 12, 24, 20, 52, 52, 18, 18, 58, 16, 14, 16, 70])
    compact_worksheet(pending_ws, row_height=22, freeze_panes="F2", zoom=85)
    add_reply_assignment_validations(pending_ws, max(len(assignment) + 20, 200))

    processed_ws = wb.create_sheet("已处理")
    append_mapped_table(processed_ws, PROCESSED_COLUMNS, processed_rows)
    set_widths(processed_ws, [52, 52, 16, 14, 16, 70, 16, 16, 70])
    compact_worksheet(processed_ws, row_height=22, freeze_panes="A2", zoom=85)

    community_ws = wb.create_sheet("社区库")
    append_mapped_table(community_ws, COMMUNITY_COLUMNS, community_rows(rows))
    set_widths(community_ws, [18, 24, 14, 14, 14, 14, 14, 14, 14, 14, 24, 70])
    compact_worksheet(community_ws, row_height=22, freeze_panes="A2", zoom=90)

    guide_ws = wb.create_sheet("字段说明")
    append_reply_assignment_field_guide(guide_ws)
    compact_worksheet(guide_ws, row_height=24, freeze_panes="A2", zoom=85)
    return assignment, processed_rows


def append_overview(ws, mode, rows, product_rows):
    counts = {
        "全部帖子": len(rows),
        "A 级帖子": sum(1 for r in rows if r.get("Priority Tier") == "A"),
        "行动队列帖子": len(action_queue_rows(rows)),
        "产品组帖子": len(product_rows),
        "P0 今天优先回复": sum(1 for r in rows if r.get("处理优先级") == "P0 今天优先回复"),
        "P1 产品组先看": sum(1 for r in rows if r.get("处理优先级") == "P1 产品组先看"),
    }
    overview_rows = [
        ("这是什么", "SUIJI Reddit 调研工作簿。先服务团队判断和行动，不要求从全量 Posts 里硬找重点。"),
        ("推荐阅读顺序", "1. Overview 2. Action Queue 3. Product Design 4. High Priority 5. Posts 6. Column Guide / Scoring Guide"),
        ("先回复 Rising 还是 New", "先看 Action Queue 里的 P0，通常是合格 Rising；New 只有在高度相关、低风险、明确求助时才提前处理。"),
        ("产品组怎么看", "产品组只需要看 Product Design：重量、过敏/敏感、佩戴舒适度相关帖子。产品组提供事实和问题，运营/创始人负责社区友好回复。"),
        ("运营/创始人怎么看", "先看 Action Queue 的 P0/P2；P4 不公开回复；高风险帖只人工复核或观察。"),
        ("研究/分析怎么看", "看 High Priority、Posts、Scoring Guide 和 Coverage，把真实痛点沉淀到周报。"),
        ("当前模式", mode),
        ("全部帖子", counts["全部帖子"]),
        ("A 级帖子", counts["A 级帖子"]),
        ("行动队列帖子", counts["行动队列帖子"]),
        ("产品组帖子", counts["产品组帖子"]),
        ("P0 今天优先回复", counts["P0 今天优先回复"]),
        ("P1 产品组先看", counts["P1 产品组先看"]),
        ("限制", "自动标签和分数是初筛。公开回复前必须打开 Reddit 原帖和 subreddit 规则人工复核。"),
    ]
    append_table(ws, ["项目", "说明"], [{"项目": a, "说明": b} for a, b in overview_rows])
    set_widths(ws, [28, 120])


def append_column_guide(ws):
    seen = set()
    guide_rows = []
    for sheet, columns in [
        ("Action Queue", ACTION_COLUMNS),
        ("Product Design", PRODUCT_DESIGN_COLUMNS),
        ("Posts", POST_COLUMNS),
        ("Daily Rising", DAILY_RISING_COLUMNS),
    ]:
        for label, _ in columns:
            key = (sheet, label)
            if key in seen:
                continue
            seen.add(key)
            definition, importance, owner, usage = COLUMN_GUIDE.get(
                label,
                ("工作簿辅助字段，用于追溯、过滤或人工复核。", "辅助", "研究/分析", "需要时结合原帖和上下文字段查看。"),
            )
            guide_rows.append({
                "所在 Sheet": sheet,
                "列名": label,
                "这列是什么": definition,
                "重要性": importance,
                "主要使用者": owner,
                "怎么用": usage,
            })
    append_table(ws, ["所在 Sheet", "列名", "这列是什么", "重要性", "主要使用者", "怎么用"], guide_rows)
    set_widths(ws, [20, 32, 70, 12, 18, 70])


def should_handle_today(row):
    if row.get("Reply Risk") == "高":
        return "否"
    if row.get("Score Band") == "A" and row.get("Reply Opportunity") in {"是", "可能"}:
        return "是"
    if row.get("Jewelry Relevance", 0) >= 4 and row.get("每小时评论数") not in ("", None) and row.get("每小时评论数") >= 1:
        return "是"
    return "观察"


def action_fields(row):
    source = row.get("Source Type") or ""
    priority = row.get("Priority Tier") or ""
    risk = row.get("Reply Risk") or ""
    reply_opp = row.get("Reply Opportunity") or ""
    is_design = row.get("是否产品组处理") == "是"
    relevance = row.get("Jewelry Relevance") or 0
    comments_per_hour = row.get("每小时评论数") or 0
    score_per_hour = row.get("每小时得分") or 0

    if risk == "高":
        return {
            "处理优先级": "P4 暂不公开回复",
            "回复顺序建议": "不建议直接回复",
            "优先理由": "回复风险高，可能涉及医疗/身体建议或版规敏感；先人工复核或只做观察。",
        }
    if is_design and priority in {"A", "B"}:
        return {
            "处理优先级": "P1 产品组先看",
            "回复顺序建议": "先产品组判断，再由运营/创始人决定是否公开回复",
            "优先理由": "命中重量、过敏/敏感或佩戴舒适度真实反馈，需要产品组先给事实判断和追问点。",
        }
    if source == "rising" and priority == "A" and reply_opp in {"是", "可能"}:
        return {
            "处理优先级": "P0 今天优先回复",
            "回复顺序建议": "先 Rising",
            "优先理由": "帖子正在上升且重要性高，适合今天优先人工复核并温和参与。",
        }
    if source == "new" and priority == "A" and reply_opp in {"是", "可能"} and relevance >= 4:
        return {
            "处理优先级": "P2 本周准备回复",
            "回复顺序建议": "Rising 后 New",
            "优先理由": "帖子很新且高度相关，但还未证明热度；可先观察互动，再决定是否回复。",
        }
    if priority in {"A", "B"} and reply_opp in {"是", "可能"}:
        return {
            "处理优先级": "P2 本周准备回复",
            "回复顺序建议": "按重要性和风险人工复核",
            "优先理由": "具备调研或互动价值，但不是今天必须抢占的 Rising 机会。",
        }
    if relevance >= 4 or comments_per_hour >= 1 or score_per_hour >= 5:
        return {
            "处理优先级": "P3 收藏观察",
            "回复顺序建议": "先观察",
            "优先理由": "有一定信号，但回复机会、意图或风险还不够明确。",
        }
    return {
        "处理优先级": "P4 暂不公开回复",
        "回复顺序建议": "不优先",
        "优先理由": "首饰相关性、痛点强度或回复机会不足，不建议占用养号和运营精力。",
    }


def is_product_design_row(row):
    return row.get("是否产品组处理") == "是"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--rules")
    parser.add_argument("--community-metadata")
    parser.add_argument("--output", required=True)
    parser.add_argument("--mode", choices=["weekly", "daily", "reply_assignment"], default="weekly")
    parser.add_argument("--master", help="Existing master workbook used to avoid already handled posts and populate 已处理.")
    parser.add_argument("--assignment-limit", type=int, default=80, help="Maximum rows in reply_assignment 待回复 sheet.")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    rules_data = load_rules(args.rules)
    community_metadata = load_community_metadata(args.community_metadata, data.get("community_metadata", {}))
    meta = data.get("meta", {})
    reference_dt = datetime.fromisoformat((meta.get("finished_at") or datetime.now(timezone.utc).isoformat()).replace("Z", "+00:00"))
    collection_date = iso_to_date(meta.get("finished_at") or meta.get("started_at"))
    rows = build_rows(data.get("posts", []), rules_data, collection_date, reference_dt, community_metadata)

    wb = Workbook()
    product_rows = [r for r in rows if is_product_design_row(r)]

    if args.mode == "reply_assignment":
        assignment, processed = build_reply_assignment_workbook(wb, rows, args.master, args.assignment_limit)
        out = Path(args.output)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(json.dumps({
            "output": str(out),
            "mode": args.mode,
            "pending_reply": len(assignment),
            "processed": len(processed),
        }, ensure_ascii=False))
        return

    overview = wb.active
    overview.title = "Overview"
    append_overview(overview, args.mode, rows, product_rows)
    compact_worksheet(overview, row_height=34, freeze_panes="A2", zoom=90, wrap_data=True)

    action_ws = wb.create_sheet("Action Queue")
    append_mapped_table(action_ws, ACTION_COLUMNS, action_queue_rows(rows))
    set_widths(action_ws, [18, 28, 54, 16, 16, 14, 14, 10, 14, 14, 18, 16, 18, 52, 52, 16, 14, 14, 14, 12, 22, 42, 36, 38, 22, 70, 60, 80, 80])
    apply_width_overrides(action_ws, {"C": 38, "N": 44, "O": 48, "Z": 44, "AA": 44, "AB": 48, "AC": 48})
    compact_worksheet(action_ws, row_height=22, freeze_panes="F2", zoom=85)

    hp = wb.create_sheet("High Priority")
    append_mapped_table(hp, ACTION_COLUMNS, [r for r in rows if r["Priority Tier"] == "A"])
    set_widths(hp, [18, 28, 54, 16, 16, 14, 14, 10, 14, 14, 18, 16, 18, 52, 52, 16, 14, 14, 14, 12, 22, 42, 36, 38, 22, 70, 60, 80, 80])
    apply_width_overrides(hp, {"C": 38, "N": 44, "O": 48, "Z": 44, "AA": 44, "AB": 48, "AC": 48})
    compact_worksheet(hp, row_height=22, freeze_panes="F2", zoom=85)

    product_ws = wb.create_sheet("Product Design")
    append_mapped_table(product_ws, PRODUCT_DESIGN_COLUMNS, product_rows)
    set_widths(product_ws, [18, 54, 18, 16, 18, 52, 52, 16, 12, 22, 22, 42, 36, 38, 16, 78, 58, 60, 80, 80])
    apply_width_overrides(product_ws, {"B": 42, "F": 42, "G": 44, "P": 48, "Q": 44, "R": 44, "S": 48, "T": 48})
    compact_worksheet(product_ws, row_height=22, freeze_panes="E2", zoom=85)

    ws = wb.create_sheet("Posts")
    append_post_table(ws, rows)
    set_widths(ws, [8, 14, 12, 12, 18, 24, 14, 14, 16, 16, 14, 16, 16, 20, 48, 46, 22, 14, 14, 12, 14, 14, 14, 16, 16, 16, 20, 36, 34, 34, 24, 16, 16, 16, 48, 48, 24, 14, 14, 12, 14, 18, 48, 48, 70, 70])
    apply_width_overrides(ws, {"Y": 44, "Z": 48, "AI": 18, "AJ": 26, "AK": 30, "AL": 30, "AM": 30, "BE": 44, "BF": 44, "BG": 46, "BH": 46, "BI": 52})
    hide_columns_by_header(ws, [
        "API订阅人数", "Sidebar订阅人数", "页面周活跃用户数", "页面周活跃用户数显示",
        "页面周贡献数", "页面周贡献数显示", "社区当前活跃人数", "社区当前活跃人数显示",
        "页面成员标签", "页面活跃标签", "社区数据来源", "社区数据状态", "搜索关键词",
    ])
    compact_worksheet(ws, row_height=18, freeze_panes="J2", zoom=80)

    if args.mode == "daily":
        rising_rows = []
        for row in rows:
            if row.get("Source Type") == "rising" or row.get("帖子年龄（小时）", 999) <= 30:
                daily = dict(row)
                daily["是否建议今天处理"] = should_handle_today(row)
                rising_rows.append(daily)
        rising_rows.sort(key=action_sort_key)
        daily_ws = wb.create_sheet("Daily Rising")
        append_mapped_table(daily_ws, DAILY_RISING_COLUMNS, rising_rows)
        set_widths(daily_ws, [18, 28, 54, 16, 16, 14, 14, 10, 14, 14, 18, 16, 18, 52, 52, 16, 14, 14, 14, 12, 22, 42, 36, 38, 22, 70, 60, 80, 80, 18])
        apply_width_overrides(daily_ws, {"C": 38, "N": 44, "O": 48, "Z": 44, "AA": 44, "AB": 48, "AC": 48})
        compact_worksheet(daily_ws, row_height=22, freeze_panes="F2", zoom=85)

    rules_ws = wb.create_sheet("Subreddit Rules")
    rules_rows = []
    for sub in sorted({r["Subreddit"] for r in rows if r.get("Subreddit")}):
        item = rules_data.get(sub) or {}
        rules = item.get("rules") or []
        rules_rows.append({
            "Subreddit": sub,
            "Rules URL": item.get("rules_url", f"https://www.reddit.com/r/{sub}/about/rules"),
            "Rules Summary": summarize_rules(sub, rules_data),
            "Rule Count": len(rules),
            "Needs Manual Review": "Yes" if not rules else "Before posting",
            "Raw Rule Names": "; ".join([x.get("short_name", "") for x in rules if x.get("short_name")]),
        })
    append_table(rules_ws, ["Subreddit", "Rules URL", "Rules Summary", "Rule Count", "Needs Manual Review", "Raw Rule Names"], rules_rows)
    set_widths(rules_ws, [22, 52, 70, 12, 22, 80])
    apply_width_overrides(rules_ws, {"B": 42, "C": 52, "F": 52})
    compact_worksheet(rules_ws, row_height=24, freeze_panes="A2", zoom=90)

    tag_ws = wb.create_sheet("Tag Dictionary")
    tag_rows = [
        ("痛点标签", "掉色/氧化/镀层磨损", "用户提到 tarnish、plating wear、氧化、变色、绿皮肤等。"),
        ("痛点标签", "敏感/刺激/发炎", "用户提到 sensitive ears、nickel、irritation、bump、allergy 等。"),
        ("痛点标签", "舒适度/重量", "用户提到太重、疼、不舒服、耳垂下垂、睡觉不舒服等。"),
        ("痛点标签", "损坏/维修/拆装困难", "用户提到断裂、卡住、松动、部件掉落、维修/修复等。"),
        ("痛点标签", "造型/职场场景", "用户在办公室、正式场合、婚礼、日常造型中不知道如何选择。"),
        ("痛点标签", "旅行/收纳/缠绕", "用户提到旅行携带、首饰缠绕、收纳、丢失、胶囊衣橱。"),
        ("评分字段", "Jewelry Relevance", "1-5 的首饰相关性门槛。相关性低的帖子即使高赞也不能进入 A 级。"),
        ("评分字段", "Reddit Score", "Reddit 原帖自己的 upvote/downvote 综合热度，不代表对 SUIJI 的调研价值。"),
        ("评分字段", "Importance Score", "SUIJI 调研价值分，0-100，综合首饰相关性、痛点、意图、产品 fit、互动量和回复机会。"),
        ("评分字段", "Score Band", "按 Importance Score 分成 A/B/C/D 区间：A=70+，B=50-69，C=30-49，D<30。"),
        ("用户意图", "抱怨/负面体验", "用户在表达问题、后悔、踩雷、差评或强烈不满。"),
        ("用户意图", "求建议/求解决方案", "用户在问怎么办、买什么、如何修复、如何选择。"),
        ("产品相关", "钛合金/低敏材质", "和 SUIJI Grade 5 titanium、低敏、敏感耳、绿皮肤等相关。"),
        ("产品相关", "Core-Lock/快拆结构", "和可拆、快换、锁扣、部件掉落、拆不下等相关。"),
        ("产品相关", "旅行/胶囊衣橱", "和少带几件、多场景切换、旅行收纳相关。"),
        ("首饰类型", "耳饰/戒指/项链/手链/穿孔饰品", "用于产品组快速判断帖子讨论的是哪类首饰或身体佩戴场景。"),
        ("首饰类型", "扣件/结构", "用户讨论锁扣、耳堵、快拆、部件松动、掉落、拆不下或可转换结构。"),
        ("首饰类型", "材质/过敏", "用户讨论钛、镍、低敏、镀层、银、金、绿皮肤、氧化或刺激问题。"),
        ("首饰类型", "佩戴舒适度", "用户讨论重量、疼痛、睡觉佩戴、松紧、摩擦、夹痛等佩戴体验。"),
        ("首饰类型", "收纳/旅行/造型/职场", "用户讨论旅行携带、收纳、胶囊衣橱、办公室或正式场景搭配。"),
        ("产品组字段", "产品组痛点类型", "专门标注重量、过敏/敏感、佩戴舒适度三类真实产品体验痛点。"),
        ("产品组字段", "是否首饰类/首饰类型", "产品组报告必须额外标注帖子是否首饰类，以及具体首饰类型，避免把泛穿搭或泛生活方式帖子误交给产品组。"),
        ("产品组字段", "负责人", "如果命中重量、过敏/敏感、佩戴舒适度真实反馈，负责人默认标为“产品组”；产品组提供事实和追问点，运营/创始人负责社区友好回复。"),
        ("社区字段", "API订阅人数", "来自 Reddit about.json 的 subscribers 字段，是 API 层面的社区订阅规模。"),
        ("社区字段", "Sidebar订阅人数", "来自 subreddit 页面 Subscriber Info 区块，例如 994,828 piercing enthusiasts subscribed。"),
        ("社区字段", "页面周活跃用户数/页面周贡献数", "来自 subreddit 页面 header 属性 weekly-active-users 与 weekly-contributions，对应页面上类似 681K / 11K 的两个可见数字。"),
        ("社区字段", "社区成员数/当前活跃人数", "优先使用页面/Sidebar 读数补充 about.json；如果抓取失败，标记为未知/待人工补充，不阻塞报告。"),
        ("社区字段", "社区规模档位/活跃度档位", "用于判断同样的帖子在多大社区和多活跃语境中出现，辅助排序和团队优先级。"),
    ]
    append_table(tag_ws, ["Tag Category", "Tag", "Definition"], [{"Tag Category": a, "Tag": b, "Definition": c} for a, b, c in tag_rows])
    set_widths(tag_ws, [22, 32, 90])
    apply_width_overrides(tag_ws, {"C": 64})
    compact_worksheet(tag_ws, row_height=24, freeze_panes="A2", zoom=90)

    column_ws = wb.create_sheet("Column Guide")
    append_column_guide(column_ws)
    compact_worksheet(column_ws, row_height=24, freeze_panes="A2", zoom=85)

    scoring_ws = wb.create_sheet("Scoring Guide")
    scoring_rows = [
        ("使用顺序", "先按“重要性分数”降序看，再用“分数区间”“回复机会”“回复风险”“首饰相关性”筛选。不要只按 Reddit 原帖分数筛。"),
        ("Reddit 原帖分数", "来自 Reddit 原始数据，表示帖子在社区里的 upvote/downvote 综合热度。它只少量影响重要性分数，高热度但不相关的帖子会被压低。"),
        ("重要性分数", "0-100 分，是 SUIJI 调研价值分。它由首饰相关性、痛点、用户意图、产品相关性、情绪强度、评论数、Reddit 原帖分数、降权/封顶规则共同决定。"),
        ("分数区间", "A=70+：重点人工看/可进报告；B=50-69：有价值，适合观察或收藏；C=30-49：弱相关，保留但不优先；D<30：低价值、泛话题或只碰巧命中关键词。"),
        ("首饰相关性规则", "基础为 0-5 分。帖子来自核心首饰/耳饰/穿孔社区加 3 分；标题/正文命中 jewelry、earring、ring、necklace、silver、gold、tarnish、piercing、titanium 等首饰词，最多加 2 分；命中搜索关键词再加 1 分；总分封顶 5。"),
        ("首饰相关性加分", "重要性分数中，首饰相关性每 1 分加 6 分，最多 30 分。"),
        ("痛点标签加分", "只要命中任意痛点标签，加 8 分；每个痛点标签再加 2 分，最多额外 9 分。痛点包括掉色/氧化、敏感/刺激、舒适度/重量、损坏/维修/拆装、造型/职场、旅行/收纳、价格/质量/信任。"),
        ("强痛点额外加分", "如果痛点不是纯“造型/职场场景”或“价格/质量/信任”，而是掉色、敏感、重量、损坏、拆装、旅行收纳等更具体痛点，额外加 6 分。"),
        ("用户意图加分", "意图为“抱怨/负面体验”加 12 分；“求建议/求解决方案”加 9 分；“购买研究/品牌比较”加 7 分。展示/晒图本身不加分。"),
        ("产品相关性加分", "每命中一个 SUIJI 产品相关标签加 2 分，最多 8 分。产品相关标签包括钛合金/低敏材质、Core-Lock/快拆结构、职场到社交切换、旅行/胶囊衣橱、质量/价值证明。"),
        ("互动量加分", "评论数每 25 条加 1 分，最多 6 分。Reddit 原帖分数每 350 分加 1 分，最多 3 分。互动量是辅助信号，不会让无关帖变成高价值。"),
        ("Rising 指标", "Daily Rising 表会计算帖子年龄、每小时得分、每小时评论数，用于判断今天是否值得快速处理。"),
        ("情绪强度加分", "情绪强度 1-5 分，每 1 级加 2 分，最多 10 分。命中 hate、cry、nightmare、disaster、regret、scam、desperate、impossible、worried、struggling、pain、hurts 等词会提高情绪强度。"),
        ("纯展示降权", "如果没有痛点且只是展示/晒图，先扣 15 分。"),
        ("无行动价值封顶", "如果没有强痛点、没有抱怨/求建议/购买研究/维修保养，也没有强产品相关信号，重要性分数最高只能到 38。"),
        ("展示帖封顶", "如果没有强痛点，只是展示/晒图，且没有强产品相关信号，重要性分数最高只能到 45。"),
        ("弱产品信号封顶", "如果没有强痛点，也没有强产品相关信号，重要性分数最高只能到 68。"),
        ("纯造型弱信号封顶", "如果唯一痛点是“造型/职场场景”，且不是求建议帖，重要性分数最高只能到 45。"),
        ("首饰相关性封顶", "首饰相关性 <=1 时，重要性分数最高 24；相关性=2 时最高 45；相关性=3 时最高 68。这样可以防止泛生活方式高赞帖排到前面。"),
        ("回复风险规则", "r/piercing、r/PiercingAdvice、r/SkincareAddiction 默认高风险，因为容易涉及医疗/身体建议；规则缺失或涉及敏感/刺激/发炎时风险至少为中。"),
        ("回复建议规则", "只有低风险或中风险帖才给温和研究式回复建议。回复建议必须结合 subreddit 规则摘要和帖子内容，避免硬广、群发、假装用户、医疗承诺。"),
        ("产品组处理规则", "命中重量、过敏/敏感、佩戴舒适度且具有抱怨、求建议、购买研究或明确痛感描述的帖子，标记“是否产品组处理=是”，负责人为“产品组”。产品组不一定直接回复 Reddit，而是提供设计判断和可问问题。"),
        ("产品组报告标注", "产品组报告必须包含“是否首饰类”和“首饰类型”。首饰类型用于区分耳饰、戒指、项链、手链、穿孔饰品、材质/过敏、扣件/结构、佩戴舒适度、收纳/旅行、造型/职场。"),
        ("社区规模规则", "成员数和当前活跃人数只作为上下文，不直接决定重要性分数。小社区的高质量真实痛点仍可进入 A/B；大社区的泛展示帖仍会被降权。"),
        ("社区数据规则", "优先保存多个来源：API订阅人数来自 about.json；Sidebar订阅人数来自 Subscriber Info；页面周活跃用户数和页面周贡献数来自 subreddit header。不同来源可能不一致，报告中必须保留来源和抓取状态。"),
        ("人工复核提醒", "标签和分数是规则自动生成，适合初筛。真正要回复前，仍需人工打开帖子和 subreddit 规则复核语境。"),
    ]
    append_table(scoring_ws, ["项目", "说明"], [{"项目": a, "说明": b} for a, b in scoring_rows])
    set_widths(scoring_ws, [28, 120])
    apply_width_overrides(scoring_ws, {"B": 76})
    compact_worksheet(scoring_ws, row_height=24, freeze_panes="A2", zoom=90)

    cov = wb.create_sheet("Coverage")
    coverage_rows = [
        ("Input File", args.input),
        ("Rules File", args.rules or ""),
        ("Generated At", datetime.now(timezone.utc).isoformat()),
        ("Collection Started", meta.get("started_at", "")),
        ("Collection Finished", meta.get("finished_at", "")),
        ("Days", meta.get("days", "")),
        ("Total Unique Posts", len(data.get("posts", []))),
        ("Errors", len(meta.get("errors", []))),
        ("Subreddits", ", ".join(meta.get("subreddits", []))),
        ("Queries", ", ".join(meta.get("queries", []))),
        ("Workbook Mode", args.mode),
        ("Community Metadata Collected", meta.get("community_metadata_collected", "")),
        ("Limitations", "Post-level analysis only unless comments are separately fetched. Subreddit rules are summaries; manually review before posting. Labels are rule-based and should be spot-checked by humans."),
    ]
    append_table(cov, ["Field", "Value"], [{"Field": a, "Value": b} for a, b in coverage_rows])
    set_widths(cov, [28, 140])
    apply_width_overrides(cov, {"B": 90})
    compact_worksheet(cov, row_height=22, freeze_panes="A2", zoom=90)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(json.dumps({
        "output": str(out),
        "posts": len(rows),
        "high_priority": sum(1 for r in rows if r["Priority Tier"] == "A"),
        "product_design": sum(1 for r in rows if is_product_design_row(r)),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
